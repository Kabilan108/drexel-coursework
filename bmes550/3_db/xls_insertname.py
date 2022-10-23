"""
Authors: [Tony Kabilan Okeke](mailto:tko35@drexel.edu),
         [Cooper Molloy](mailto:cdm348@drexel.edu)  

This function will:
- Create the excel file if it doesn't exist and populate a header row 
  with the columns 'id' and 'name'
- Insert the given name into the excel file and auto-increment the id
- Return the number of names in the table
"""

from openpyxl import Workbook, load_workbook
import os

def xls_insertname(xlsfile: str, name: str) -> int:

    # Check if excel file exists
    if not os.path.exists(xlsfile):
        wb = Workbook()
        ws = wb.active
        ws.append(['id', 'name'])
        wb.save(xlsfile)
    else:
        wb = load_workbook(xlsfile)
        ws = wb.active

    # Get the last row number
    last_row = ws.max_row

    # Insert the name
    ws.append([last_row, name])

    # Save and close the workbook
    wb.save(xlsfile)
    wb.close()

    return last_row
